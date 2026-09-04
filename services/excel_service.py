from __future__ import annotations

import logging
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook


LOGGER = logging.getLogger(__name__)


class ExcelService:
    """Small replaceable persistence adapter for the MVP."""

    def __init__(self, path: Path | None = None) -> None:
        default = Path.home() / ".flashtile" / "FlashTile.xlsx"
        self.path = path or default
        self.path.parent.mkdir(parents=True, exist_ok=True)
        if not self.path.exists():
            self._create()
        try:
            self._ensure_schema()
        except Exception:
            self._preserve_corrupt_workbook()
            self._create()
            self._ensure_schema()

    def _create(self) -> None:
        workbook = Workbook()
        progress = workbook.active
        progress.title = "Progress"
        progress.append(["date", "topic", "lesson", "xp"])
        settings = workbook.create_sheet("Settings")
        settings.append(["key", "value"])
        settings.append(["selected_topic", "AWS & Cloud"])
        settings.append(["welcome_seen_rc15", False])
        bookmarks = workbook.create_sheet("Bookmarks")
        bookmarks.append(["topic", "lesson", "saved_on"])
        notes = workbook.create_sheet("Notes")
        notes.append(["topic", "lesson", "note", "updated_on"])
        reviews = workbook.create_sheet("Reviews")
        reviews.append(
            [
                "topic",
                "lesson",
                "mastery_level",
                "due_date",
                "interval_days",
                "confidence",
                "last_result",
                "updated_on",
            ]
        )
        workbook.save(self.path)

    def _preserve_corrupt_workbook(self) -> None:
        if not self.path.exists():
            return
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        preserved = self.path.with_name(
            f"{self.path.stem}.corrupt-{timestamp}{self.path.suffix}"
        )
        shutil.move(str(self.path), str(preserved))
        LOGGER.exception("Preserved unreadable workbook at %s", preserved)

    def _backup(self) -> None:
        if not self.path.exists():
            return
        backup_dir = self.path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        backup = backup_dir / f"{self.path.stem}-{date.today().isoformat()}.xlsx"
        if not backup.exists():
            shutil.copy2(self.path, backup)

    def reset_demo_data(self) -> Path:
        """Back up the current workbook, then restore a clean demo state."""
        backup_dir = self.path.parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        backup = backup_dir / f"{self.path.stem}-demo-reset-{timestamp}.xlsx"
        if self.path.exists():
            shutil.copy2(self.path, backup)
        self._create()
        return backup

    def _save(self, workbook) -> None:
        self._backup()
        temporary = self.path.with_suffix(".tmp.xlsx")
        workbook.save(temporary)
        temporary.replace(self.path)

    def _ensure_schema(self) -> None:
        """Add newer sheets without replacing an existing user workbook."""
        workbook = load_workbook(self.path)
        changed = False
        if "Bookmarks" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Bookmarks")
            sheet.append(["topic", "lesson", "saved_on"])
            changed = True
        if "Notes" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Notes")
            sheet.append(["topic", "lesson", "note", "updated_on"])
            changed = True
        if "Reviews" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Reviews")
            sheet.append(
                [
                    "topic",
                    "lesson",
                    "mastery_level",
                    "due_date",
                    "interval_days",
                    "confidence",
                    "last_result",
                    "updated_on",
                ]
            )
            changed = True
        if changed:
            self._save(workbook)

    def selected_topic(self) -> str:
        return str(self._setting("selected_topic", "AWS & Cloud"))

    def _setting(self, key: str, default: object) -> object:
        workbook = load_workbook(self.path)
        for stored_key, value in workbook["Settings"].iter_rows(min_row=2, values_only=True):
            if stored_key == key:
                return value
        return default

    def _set_setting(self, key: str, value: object) -> None:
        workbook = load_workbook(self.path)
        sheet = workbook["Settings"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == key:
                sheet.cell(row, 2).value = value
                break
        else:
            sheet.append([key, value])
        self._save(workbook)

    def set_selected_topic(self, topic: str) -> None:
        self._set_setting("selected_topic", topic)

    def learning_goal(self) -> str:
        return str(self._setting("learning_goal", "") or "")

    def set_learning_goal(self, goal: str) -> None:
        self._set_setting("learning_goal", goal)

    def learning_goal_position(self) -> int:
        return int(self._setting("learning_goal_position", 0) or 0)

    def set_learning_goal_position(self, position: int) -> None:
        self._set_setting("learning_goal_position", max(0, position))

    def reduced_motion(self) -> bool:
        value = self._setting("reduced_motion", False)
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "on"}

    def set_reduced_motion(self, enabled: bool) -> None:
        self._set_setting("reduced_motion", bool(enabled))

    def welcome_seen(self) -> bool:
        value = self._setting("welcome_seen_rc15", False)
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {"1", "true", "yes", "on"}

    def set_welcome_seen(self, seen: bool) -> None:
        self._set_setting("welcome_seen_rc15", bool(seen))

    def lesson_index(self, topic: str) -> int:
        return int(self._setting(f"lesson_index::{topic}", 0) or 0)

    def set_lesson_index(self, topic: str, index: int) -> None:
        self._set_setting(f"lesson_index::{topic}", index)

    def is_bookmarked(self, topic: str, lesson: str) -> bool:
        workbook = load_workbook(self.path)
        return any(
            row[0] == topic and row[1] == lesson
            for row in workbook["Bookmarks"].iter_rows(min_row=2, values_only=True)
        )

    def set_bookmarked(self, topic: str, lesson: str, bookmarked: bool) -> None:
        workbook = load_workbook(self.path)
        sheet = workbook["Bookmarks"]
        matching_row = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == topic and sheet.cell(row, 2).value == lesson:
                matching_row = row
                break
        if bookmarked and matching_row is None:
            sheet.append([topic, lesson, date.today().isoformat()])
        elif not bookmarked and matching_row is not None:
            sheet.delete_rows(matching_row)
        self._save(workbook)

    def bookmarks(self) -> list[tuple[str, str]]:
        workbook = load_workbook(self.path)
        return [
            (str(row[0]), str(row[1]))
            for row in workbook["Bookmarks"].iter_rows(min_row=2, values_only=True)
            if row[0] and row[1]
        ]

    def lesson_note(self, topic: str, lesson: str) -> str:
        workbook = load_workbook(self.path)
        for stored_topic, stored_lesson, note, _ in workbook["Notes"].iter_rows(
            min_row=2, values_only=True
        ):
            if stored_topic == topic and stored_lesson == lesson:
                return str(note or "")
        return ""

    def save_lesson_note(self, topic: str, lesson: str, note: str) -> None:
        clean_note = note.strip()[:500]
        workbook = load_workbook(self.path)
        sheet = workbook["Notes"]
        matching_row = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == topic and sheet.cell(row, 2).value == lesson:
                matching_row = row
                break
        if clean_note:
            values = [topic, lesson, clean_note, date.today().isoformat()]
            if matching_row is None:
                sheet.append(values)
            else:
                for column, value in enumerate(values, start=1):
                    sheet.cell(matching_row, column).value = value
        elif matching_row is not None:
            sheet.delete_rows(matching_row)
        self._save(workbook)

    def notes(self) -> list[tuple[str, str, str]]:
        workbook = load_workbook(self.path)
        rows = list(workbook["Notes"].iter_rows(min_row=2, values_only=True))
        return [
            (str(row[0]), str(row[1]), str(row[2]))
            for row in reversed(rows)
            if row[0] and row[1] and row[2]
        ]

    def review_state(self, topic: str, lesson: str) -> dict[str, object]:
        workbook = load_workbook(self.path)
        for row in workbook["Reviews"].iter_rows(min_row=2, values_only=True):
            if row[0] == topic and row[1] == lesson:
                return {
                    "mastery_level": int(row[2] or 0),
                    "due_date": str(row[3] or ""),
                    "interval_days": int(row[4] or 0),
                    "confidence": str(row[5] or ""),
                    "last_result": str(row[6] or ""),
                }
        return {
            "mastery_level": 0,
            "due_date": "",
            "interval_days": 0,
            "confidence": "",
            "last_result": "",
        }

    def _save_review_state(
        self,
        topic: str,
        lesson: str,
        mastery_level: int,
        due_date: str,
        interval_days: int,
        confidence: str,
        last_result: str,
    ) -> None:
        workbook = load_workbook(self.path)
        sheet = workbook["Reviews"]
        matching_row = None
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == topic and sheet.cell(row, 2).value == lesson:
                matching_row = row
                break
        values = [
            topic,
            lesson,
            max(0, min(3, mastery_level)),
            due_date,
            max(0, interval_days),
            confidence,
            last_result,
            date.today().isoformat(),
        ]
        if matching_row is None:
            sheet.append(values)
        else:
            for column, value in enumerate(values, start=1):
                sheet.cell(matching_row, column).value = value
        self._save(workbook)

    def set_confidence(self, topic: str, lesson: str, confidence: str) -> None:
        state = self.review_state(topic, lesson)
        if confidence == "got_it":
            level, interval = max(2, int(state["mastery_level"])), 3
        elif confidence == "need_practice":
            level, interval = 1, 1
        else:
            level, interval = max(1, int(state["mastery_level"])), 3
        due = (date.today() + timedelta(days=interval)).isoformat()
        self._save_review_state(
            topic, lesson, level, due, interval, confidence, "lesson_complete"
        )

    def record_review_result(self, topic: str, lesson: str, correct: bool) -> None:
        state = self.review_state(topic, lesson)
        current_level = int(state["mastery_level"])
        if correct:
            level = min(3, max(1, current_level + 1))
            interval = {1: 3, 2: 7, 3: 14}[level]
            result = "correct"
        else:
            level = max(1, current_level - 1)
            interval = 1
            result = "incorrect"
        due = (date.today() + timedelta(days=interval)).isoformat()
        self._save_review_state(
            topic,
            lesson,
            level,
            due,
            interval,
            str(state["confidence"]),
            result,
        )

    def due_reviews(self) -> list[tuple[str, str]]:
        workbook = load_workbook(self.path)
        today = date.today().isoformat()
        due = []
        for row in workbook["Reviews"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[3] and str(row[3]) <= today:
                due.append((str(row[0]), str(row[1])))
        return due

    def complete(self, topic: str, lesson: str, xp: int) -> bool:
        workbook = load_workbook(self.path)
        sheet = workbook["Progress"]
        today = date.today().isoformat()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == today and row[1] == topic and row[2] == lesson:
                return False
        sheet.append([today, topic, lesson, xp])
        self._save(workbook)
        return True

    def totals(self) -> tuple[int, int]:
        workbook = load_workbook(self.path)
        rows = list(workbook["Progress"].iter_rows(min_row=2, values_only=True))
        xp = sum(int(row[3] or 0) for row in rows)
        unique_days = sorted({str(row[0]) for row in rows if row[0]}, reverse=True)
        streak = 0
        cursor = date.today()
        for value in unique_days:
            if value == cursor.isoformat():
                streak += 1
                cursor = date.fromordinal(cursor.toordinal() - 1)
            elif value < cursor.isoformat():
                break
        return xp, streak

    def progress_summary(self, total_lessons: int) -> dict[str, object]:
        workbook = load_workbook(self.path)
        completed = {
            (str(row[1]), str(row[2]))
            for row in workbook["Progress"].iter_rows(min_row=2, values_only=True)
            if row[1] and row[2] and not str(row[2]).endswith(" • Recall")
        }
        completed_topics = {topic for topic, _ in completed}
        mastered = sum(
            1
            for row in workbook["Reviews"].iter_rows(min_row=2, values_only=True)
            if row[0] and row[1] and int(row[2] or 0) >= 3
        )
        note_count = sum(
            1
            for row in workbook["Notes"].iter_rows(min_row=2, values_only=True)
            if row[0] and row[1] and row[2]
        )
        bookmark_count = sum(
            1
            for row in workbook["Bookmarks"].iter_rows(min_row=2, values_only=True)
            if row[0] and row[1]
        )
        topic_counts: dict[str, int] = {}
        for topic, _ in completed:
            topic_counts[topic] = topic_counts.get(topic, 0) + 1
        completed_count = len(completed)
        return {
            "completed": completed_count,
            "completed_topics": len(completed_topics),
            "mastered": mastered,
            "notes": note_count,
            "bookmarks": bookmark_count,
            "due_reviews": len(self.due_reviews()),
            "percent": round(completed_count * 100 / max(1, total_lessons)),
            "topic_counts": topic_counts,
        }
