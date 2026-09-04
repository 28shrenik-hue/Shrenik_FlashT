from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from tempfile import TemporaryDirectory


ROOT = Path(__file__).resolve().parent


def report(label: str, action) -> None:
    try:
        action()
    except Exception as exc:
        print(f"FAIL  {label}: {exc}")
        raise
    print(f"PASS  {label}")


def check_python() -> None:
    if sys.version_info < (3, 9):
        raise RuntimeError("Python 3.9 or newer is required")


def check_files() -> None:
    required = [
        ROOT / "main.py",
        ROOT / "ui" / "qml" / "Main.qml",
        ROOT / "assets" / "branding" / "FlashTile_3D_Logo.png",
        ROOT / "assets" / "branding" / "FlashTile.ico",
    ]
    missing = [str(path.relative_to(ROOT)) for path in required if not path.is_file()]
    if missing:
        raise RuntimeError("Missing: " + ", ".join(missing))
    if (ROOT / "assets" / "branding" / "FlashTile_3D_Logo.png").read_bytes()[:8] != (
        b"\x89PNG\r\n\x1a\n"
    ):
        raise RuntimeError("FlashTile logo is not a valid PNG asset")


def check_data_and_learning() -> None:
    from openpyxl import load_workbook

    from services import learning_service
    from services.excel_service import ExcelService

    with TemporaryDirectory() as folder:
        workbook_path = Path(folder) / "FlashTile.xlsx"
        store = ExcelService(workbook_path)
        store.set_bookmarked("AWS & Cloud", "AWS Global Infrastructure", True)
        store.save_lesson_note(
            "AWS & Cloud",
            "AWS Global Infrastructure",
            "Availability Zones isolate failures.",
        )
        store.set_confidence(
            "AWS & Cloud", "AWS Global Infrastructure", "need_practice"
        )
        store.set_learning_goal("Balanced digital foundations")
        store.set_reduced_motion(True)

        workbook = load_workbook(workbook_path)
        expected = {
            "Progress",
            "Settings",
            "Bookmarks",
            "Notes",
            "Reviews",
            "Topic Requests",
        }
        if not expected.issubset(workbook.sheetnames):
            raise RuntimeError("Workbook schema is incomplete")
        if not (Path(folder) / "backups").exists():
            raise RuntimeError("Workbook backup was not created")

        original_store = learning_service.ExcelService
        learning_service.ExcelService = lambda: ExcelService(workbook_path)
        try:
            learning = learning_service.LearningService()
            if len(learning.topics) != 5:
                raise RuntimeError("Expected three flagship and two prepared learning areas")
            if sum(len(lessons) for lessons in learning_service.LESSONS.values()) != 25:
                raise RuntimeError("Expected twenty-five curated and prepared lessons")
            if len(learning.learningGoals) != 3:
                raise RuntimeError("Expected three category-aligned learning paths")
            if len(learning.topLearningItems) != 9:
                raise RuntimeError("Expected three flagship areas and six prepared subjects")
            if learning.learningGoal != "Build resilient cloud skills":
                raise RuntimeError("Legacy mixed-topic goal was not migrated safely")
            if not learning.reducedMotion:
                raise RuntimeError("Reduced-motion preference did not persist")
            if not all(
                [
                    learning.dailyDiscoveryCategory,
                    learning.dailyDiscoveryTitle,
                    learning.dailyDiscoveryBody,
                    learning.dailyDiscoveryContext,
                    learning.dailyDiscoverySource,
                    learning.dailyDiscoverySourceUrl,
                ]
            ):
                raise RuntimeError("Daily Discovery content is incomplete")
            if not all(
                [
                    learning.teamName,
                    learning.teamWeeklyGoal,
                    learning.teamXp,
                    learning.teamStreak,
                    len(learning.teamMemberItems) == 6,
                    learning.teamChallenge,
                ]
            ):
                raise RuntimeError("Team Board data is incomplete")
            if (
                len(learning.badgeItems) != 6
                or learning.searchResultCount != len(learning_service.LESSONS[learning.topic])
            ):
                raise RuntimeError("Progress, badges, or lesson search is incomplete")
            if learning.tourCount != 7 or not learning.tourTitle:
                raise RuntimeError("Guided Tour is incomplete")
            for lessons in learning_service.LESSONS.values():
                for lesson in lessons:
                    if not all(
                        [
                            lesson.description,
                            lesson.scenario,
                            lesson.question,
                            lesson.recall_scenario,
                            lesson.recall_question,
                        ]
                    ):
                        raise RuntimeError(f"Incomplete lesson: {lesson.title}")
        finally:
            learning_service.ExcelService = original_store


def check_qml() -> None:
    os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")
    os.environ.setdefault("QT_QUICK_BACKEND", "software")
    os.environ.setdefault("QT_QUICK_CONTROLS_STYLE", "Basic")

    from PySide6.QtCore import QMetaObject, QObject, QUrl
    from PySide6.QtGui import QGuiApplication
    from PySide6.QtQml import QQmlApplicationEngine

    from services import learning_service
    from services.excel_service import ExcelService

    with TemporaryDirectory() as folder:
        original_store = learning_service.ExcelService
        workbook_path = Path(folder) / "FlashTile.xlsx"
        ExcelService(workbook_path).set_welcome_seen(True)
        learning_service.ExcelService = lambda: ExcelService(
            workbook_path
        )
        try:
            app = QGuiApplication.instance() or QGuiApplication([])
            engine = QQmlApplicationEngine()
            learning = learning_service.LearningService()
            engine.rootContext().setContextProperty("learningService", learning)
            engine.load(
                QUrl.fromLocalFile(str(ROOT / "ui" / "qml" / "Main.qml"))
            )
            if not engine.rootObjects():
                raise RuntimeError("QML did not create a root window")
            window = engine.rootObjects()[0]
            if (window.width(), window.height()) != (410, 690):
                raise RuntimeError(
                    f"Unexpected tile size: {window.width()}x{window.height()}"
                )
            app.processEvents()
            welcome = window.findChild(QObject, "welcomeLayer")
            continue_button = window.findChild(QObject, "welcomeContinueButton")
            start_button = window.findChild(QObject, "startLearningButton")
            custom_input = window.findChild(QObject, "customGoalInput")
            browse_button = window.findChild(QObject, "browsePreparedButton")
            prepared_popup = window.findChild(QObject, "preparedTopicsPopup")
            main_drag = window.findChild(QObject, "mainHeaderDrag")
            title_drag = window.findChild(QObject, "titleHeaderDrag")
            welcome_drag = window.findChild(QObject, "welcomeHeaderDrag")
            if any(
                item is None
                for item in (
                    welcome,
                    continue_button,
                    start_button,
                    custom_input,
                    browse_button,
                    prepared_popup,
                    main_drag,
                    title_drag,
                    welcome_drag,
                )
            ):
                raise RuntimeError("Welcome Tile controls were not created")
            if not bool(welcome.property("visible")):
                raise RuntimeError("Welcome Tile was not shown on application launch")
            if not QMetaObject.invokeMethod(continue_button, "click"):
                raise RuntimeError("Learning Goal action could not be invoked")
            app.processEvents()
            if int(window.property("welcomeStep")) != 1:
                raise RuntimeError("Welcome Tile did not open Learning Goals")
            if not QMetaObject.invokeMethod(start_button, "click"):
                raise RuntimeError("Begin Learning action could not be invoked")
            app.processEvents()
            if bool(welcome.property("visible")) or not learning.welcomeSeen:
                raise RuntimeError("Start Learning did not open the learning experience")
            mapped_goals = {
                "Build resilient cloud skills": (
                    "AWS & Cloud",
                    "AWS Global Infrastructure",
                ),
                "Use AI responsibly": ("AI / ML", "How Models Learn"),
                "Strengthen digital trust": (
                    "Cybersecurity & Digital Trust",
                    "The Principle of Least Privilege",
                ),
            }
            for goal, (expected_topic, expected_title) in mapped_goals.items():
                window.setProperty("onboardingGoal", goal)
                if not QMetaObject.invokeMethod(start_button, "click"):
                    raise RuntimeError(f"Could not begin mapped goal: {goal}")
                app.processEvents()
                if learning.topic != expected_topic or learning.title != expected_title:
                    raise RuntimeError(
                        f"Goal mapping failed for {goal}: "
                        f"{learning.topic} / {learning.title}"
                    )
        finally:
            learning_service.ExcelService = original_store


def main() -> int:
    parser = argparse.ArgumentParser(description="FlashTile release self-check")
    parser.add_argument(
        "--quick",
        action="store_true",
        help="Skip offscreen QML rendering for a faster startup check",
    )
    args = parser.parse_args()

    report("Python runtime", check_python)
    report("Required files and branding", check_files)
    report("Workbook, backup, and learning services", check_data_and_learning)
    if not args.quick:
        report("QML load and compact 410x690 geometry", check_qml)
    print("FlashTile self-check complete.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
